import openpyxl

book = openpyxl.load_workbook("C:\\Users\\deepa\\OneDrive\\Desktop\\ValidLoginCredentials.xlsx")
sheet = book.active
Dict = {}
#print(sheet.cell(row=2, column=3).value)
#print(sheet.max_row)
#print(sheet.max_column)
#print(sheet["C2"].value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "testcase1":
        for j in range(2, sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)