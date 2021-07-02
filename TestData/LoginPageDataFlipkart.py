import openpyxl
class LoginPageDataFlipkart:

    test_ValidloginCredentials =[{'EnterEmailOrMobNumber': 8130253170, 'EnterPassword': "Bilaspur@123"}]
    test_InvalidloginCredentials = [{"EnterEmailOrMobNumber":8130253170, "EnterPassword":"abc@123"},{"EnterEmailOrMobNumber":8130253170, "EnterPassword":"Bipur@123"},{"EnterEmailOrMobNumber":8130253170, "EnterPassword":"abc@3"}]
    test_NumberForgetPassword = [{"EnterEmailOrMobNumber":8130253170}]

    @staticmethod
    def getTestValidCredentials(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\deepa\\OneDrive\\Desktop\\ValidLoginCredentials.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]

    @staticmethod
    def getTestInvalidCredentials():
        Dict1 = {}
        Dict2 = {}
        Dict3 = {}
        TestList = []
        book1 = openpyxl.load_workbook("C:\\Users\\deepa\\OneDrive\\Desktop\\InvalidCredentials.xlsx")
        sheet1 = book1.active
        for a in range(1, sheet1.max_row + 1):
            for b in range(2, sheet1.max_column + 1):
                Dict1[sheet1.cell(row=1, column=b).value] = sheet1.cell(row=2, column=b).value

        for c in range(1, sheet1.max_row + 1):
            for d in range(2, sheet1.max_column + 1):
                Dict2[sheet1.cell(row=1, column=d).value] = sheet1.cell(row=3, column=d).value

        for e in range(1, sheet1.max_row + 1):
            for f in range(2, sheet1.max_column + 1):
                Dict3[sheet1.cell(row=1, column=f).value] = sheet1.cell(row=4, column=f).value

        TestList.append(Dict1)
        TestList.append(Dict2)
        TestList.append(Dict3)
        return TestList